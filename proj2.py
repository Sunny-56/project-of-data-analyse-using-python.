
# import necessary libraries
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side  # importing some styles
from openpyxl.styles import PatternFill
import streamlit as st
from streamlit_option_menu import option_menu

with st.sidebar:
    selected = option_menu(
        menu_title="Main Menu",
        options=["Home", "Upload File", "Give Path"],
        icons=["house", "upload", "folder"],
        menu_icon="cast",
        default_index=0,
    )

# start_time = datetime.now()
if selected == "Home":
    st.title("Welcome to CS384 Project")
    st.subheader("Developed by Amrit Kumar (2001CB06) and Sunny Kumar (2001CB56)")

if selected == "Give Path":
    data = os.listdir('input')



# proj_octant_gui definition
def proj_octant_gui(uploaded_files, mod_value):
    # if selected == "Upload File":
    #     data = uploaded_files
    df = pd.DataFrame()
    for file in uploaded_files:
        if selected == "Upload File":
            wb = load_workbook(file)
            sheet1 = wb.active
            n = sheet1.max_row
            temp = pd.read_excel(file)
        
        if selected == "Give Path":
            path = 'input' + '/' + file
            wb = load_workbook(path)
            sheet1 = wb.active
            n = sheet1.max_row
            temp = pd.read_excel(path)

        df = df.append(temp, ignore_index=True)

        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        # list for storing u', v', w' and time
        l1, l2, l3, time = ([] for i in range(4))

        sheet1.cell(row=1, column=5).value = "U avg"
        sheet1.cell(row=1, column=6).value = "V avg"
        sheet1.cell(row=1, column=7).value = "W avg"
        sheet1.cell(row=1, column=8).value = "U'=U-U avg"
        sheet1.cell(row=1, column=9).value = "V'=V-V avg"
        sheet1.cell(row=1, column=10).value = "W=W-W avg"

        Uavg = df['U'].mean()
        Vavg = df['V'].mean()
        Wavg = df['W'].mean()

        # storing U average, V average and W average
        sheet1['E2'] = Uavg
        sheet1['F2'] = Vavg
        sheet1['G2'] = Wavg

        # storing U'= U-Uavg, V'= V-Vavg and W'= W-Wavg and inserting the values in excel
        i = 2
        while (i <= n):
            sheet1.cell(row=i, column=8).value = round(sheet1.cell(
                row=i, column=2).value - Uavg, 3)
            l1.append(sheet1.cell(row=i, column=2).value - Uavg)
            sheet1.cell(row=i, column=9).value = round(sheet1.cell(
                row=i, column=3).value - Vavg, 3)
            l2.append(sheet1.cell(row=i, column=3).value - Vavg)
            sheet1.cell(row=i, column=10).value = round(sheet1.cell(
                row=i, column=4).value - Wavg, 3)
            l3.append(sheet1.cell(row=i, column=4).value - Wavg)
            time.append(sheet1.cell(row=i, column=1).value)
            i = i+1

        sheet1.cell(row=1, column=11).value = "Octant"
        octant_list = []
        len5 = 0

        # code for counting the no of overall octant count and storing octant ids in octant_list
        i = 0
        overall_count1 = overall_count2 = overall_count3 = overall_count4 = overall_countneg1 = overall_countneg2 = overall_countneg3 = overall_countneg4 = 0
        while (i < len(l1)):
            if (l1[i] > 0 and l2[i] > 0):
                if (l3[i] > 0):
                    overall_count1 += 1
                    octant_list.insert(len5, +1)
                    len5 += 1
                else:
                    overall_countneg1 += 1
                    octant_list.insert(len5, -1)
                    len5 += 1
            elif (l1[i] < 0 and l2[i] > 0):
                if (l3[i] > 0):
                    overall_count2 += 1
                    octant_list.insert(len5, +2)
                    len5 += 1
                else:
                    overall_countneg2 += 1
                    octant_list.insert(len5, -2)
                    len5 += 1
            elif (l1[i] < 0 and l2[i] < 0):
                if (l3[i] > 0):
                    overall_count3 += 1
                    octant_list.insert(len5, +3)
                    len5 += 1
                else:
                    overall_countneg3 += 1
                    octant_list.insert(len5, -3)
                    len5 += 1
            elif (l1[i] > 0 and l2[i] < 0):
                if (l3[i] > 0):
                    overall_count4 += 1
                    octant_list.insert(len5, +4)
                    len5 += 1
                else:
                    overall_countneg4 += 1
                    octant_list.insert(len5, -4)
                    len5 += 1
            i = i+1

        # inserting the octant ids in excel file
        i = 2
        while (i <= len(octant_list)+1):
            sheet1.cell(row=i, column=11).value = octant_list[i-2]
            i = i+1

        octant_name_id_mapping = {"1": "Internal outward interaction", "-1": "External outward interaction", "2": "External Ejection",
                                "-2": "Internal Ejection", "3": "External inward interaction", "-3": "Internal inward interaction", "4": "Internal sweep", "-4": "External sweep"}
        # list for storing Rank1 Octant ID for Mod value
        count_Rank1_ModValue = []
        # funtion definition for counting no of octants in intervals

        def octant_range_names(mod):
            j = k = 0
            interval = mod

            while (j < times+1):

                count1 = count2 = count3 = count4 = countneg1 = countneg2 = countneg3 = countneg4 = 0
                while (k < mod and k < len(l1)):
                    if (l1[k] > 0 and l2[k] > 0):
                        if (l3[k] > 0):
                            count1 += 1
                        else:
                            countneg1 += 1
                    elif (l1[k] < 0 and l2[k] > 0):
                        if (l3[k] > 0):
                            count2 += 1
                        else:
                            countneg2 += 1
                    elif (l1[k] < 0 and l2[k] < 0):
                        if (l3[k] > 0):
                            count3 += 1
                        else:
                            countneg3 += 1
                    elif (l1[k] > 0 and l2[k] < 0):
                        if (l3[k] > 0):
                            count4 += 1
                        else:
                            countneg4 += 1
                    k = k+1

                mod = mod+interval
                # appending the values of octant count in val1, val_1, val2, val_2, val3, val_3, val4 and val_4 lists for all intervals
                val1.append(count1)
                val_1.append(countneg1)
                val2.append(count2)
                val_2.append(countneg2)
                val3.append(count3)
                val_3.append(countneg3)
                val4.append(count4)
                val_4.append(countneg4)
                j = j+1

                mod_list = [count1, countneg1, count2, countneg2,
                            count3, countneg3, count4, countneg4]
                list_sorted = [count1, countneg1, count2,
                            countneg2, count3, countneg3, count4, countneg4]
                list_sorted.sort(reverse=True)
                # finding and storing rank for mod interval
                sheet1.cell(row=j+4, column=23).value = list_sorted.index(count1)+1
                sheet1.cell(
                    row=j+4, column=24).value = list_sorted.index(countneg1)+1
                sheet1.cell(row=j+4, column=25).value = list_sorted.index(count2)+1
                sheet1.cell(
                    row=j+4, column=26).value = list_sorted.index(countneg2)+1
                sheet1.cell(row=j+4, column=27).value = list_sorted.index(count3)+1
                sheet1.cell(
                    row=j+4, column=28).value = list_sorted.index(countneg3)+1
                sheet1.cell(row=j+4, column=29).value = list_sorted.index(count4)+1
                sheet1.cell(
                    row=j+4, column=30).value = list_sorted.index(countneg4)+1

                for i in range(23, 31):
                    sheet1.cell(row=j+4, column=i).border = thin_border
                    if sheet1.cell(row=j+4, column=i).value == 1:
                        sheet1.cell(row=j+4, column=i).fill = PatternFill(
                            start_color='FFD970', end_color='FFD970', fill_type="solid")

                # finding and storing rank1 Octant ID and rank1 Octant Name for mod interval
                if (list_sorted.index(count1)+1 == 1):
                    sheet1.cell(row=j+4, column=31).value = 1
                    sheet1.cell(
                        row=j+4, column=32).value = octant_name_id_mapping["1"]
                elif (list_sorted.index(countneg1)+1 == 1):
                    sheet1.cell(row=j+4, column=31).value = -1
                    sheet1.cell(
                        row=j+4, column=32).value = octant_name_id_mapping["-1"]
                elif (list_sorted.index(count2)+1 == 1):
                    sheet1.cell(row=j+4, column=31).value = 2
                    sheet1.cell(
                        row=j+4, column=32).value = octant_name_id_mapping["2"]
                elif (list_sorted.index(countneg2)+1 == 1):
                    sheet1.cell(row=j+4, column=31).value = -2
                    sheet1.cell(
                        row=j+4, column=32).value = octant_name_id_mapping["-2"]
                elif (list_sorted.index(count3)+1 == 1):
                    sheet1.cell(row=j+4, column=31).value = 3
                    sheet1.cell(
                        row=j+4, column=32).value = octant_name_id_mapping["3"]
                elif (list_sorted.index(countneg3)+1 == 1):
                    sheet1.cell(row=j+4, column=31).value = -3
                    sheet1.cell(
                        row=j+4, column=32).value = octant_name_id_mapping["-3"]
                elif (list_sorted.index(count4)+1 == 1):
                    sheet1.cell(row=j+4, column=31).value = 4
                    sheet1.cell(
                        row=j+4, column=32).value = octant_name_id_mapping["4"]
                elif (list_sorted.index(countneg4)+1 == 1):
                    sheet1.cell(row=j+4, column=31).value = -4
                    sheet1.cell(
                        row=j+4, column=32).value = octant_name_id_mapping["-4"]

                sheet1.cell(row=j+4, column=31).border = thin_border
                sheet1.cell(row=j+4, column=32).border = thin_border

                count_Rank1_ModValue.append(sheet1.cell(row=j+4, column=31).value)

        # value of mod
        mod = mod_value
        ltemp = []
        ltemp.insert(0, "Overall Count")
        strr = "Mod "
        strr += str(mod)
        ltemp.insert(1, strr)
        num = n
        times = int(num/mod)
        modulo_val = num % mod
        len6 = 2
        for i in range(times):
            strr = str(mod*i)
            strr += "-"
            strr += str(mod*(i+1)-1)
            ltemp.insert(len6, strr)
            len6 += 1
        if (modulo_val):
            strr = str(num-modulo_val)
            strr += "-"
            strr += str(num)
            ltemp.insert(len6, strr)
            len6 += 1
        sheet1.cell(row=3, column=13).value = "Mod" + str(mod)
        sheet1.cell(row=2, column=14).value = "octant ID"
        sheet1.cell(row=2, column=14).border = thin_border
        i = 3
        while (i <= len(ltemp)+2):
            sheet1.cell(row=i, column=14).value = ltemp[i-3]
            sheet1.cell(row=i, column=14).border = thin_border
            i = i+1

        oct_list = ["1", "-1", "2", "-2", "3", "-3", "4", "-4"]
        Rank_list = ["Rank Octant 1", "Rank Octant -1", "Rank Octant 2", "Rank Octant -2",
                    "Rank Octant 3", "Rank Octant -3", "Rank Octant 4", "Rank Octant -4"]

        # inserting octant ids in different columns in excel file
        i = 0
        j = 14
        while (i < len(oct_list)):
            sheet1.cell(row=2, column=j+1).value = oct_list[i]
            sheet1.cell(row=2, column=j+1).border = thin_border
            sheet1.cell(row=2, column=j+9).value = Rank_list[i]
            sheet1.cell(row=2, column=j+9).border = thin_border
            j = j+1
            i = i+1
        sheet1.cell(row=2, column=31).value = "Rank1 Octant ID"
        sheet1.cell(row=2, column=31).border = thin_border
        sheet1.cell(row=2, column=32).value = "Rank1 Octant Name"
        sheet1.cell(row=2, column=32).border = thin_border

        # creating lists for storing octant count
        val1, val_1, val2, val_2, val3, val_3, val4, val_4 = ([] for i in range(8))

        # storing overall count of octants
        val1.insert(0, overall_count1)
        val_1.insert(0, overall_countneg1)
        val2.insert(0, overall_count2)
        val_2.insert(0, overall_countneg2)
        val3.insert(0, overall_count3)
        val_3.insert(0, overall_countneg3)
        val4.insert(0, overall_count4)
        val_4.insert(0, overall_countneg4)
        # leaving a box blank
        val1.insert(1, None)
        val_1.insert(1, None)
        val2.insert(1, None)
        val_2.insert(1, None)
        val3.insert(1, None)
        val_3.insert(1, None)
        val4.insert(1, None)
        val_4.insert(1, None)

        # function callling
        octant_range_names(mod)

        # inserting overall count and count in interval in sheet1
        i = 3
        while (i <= len(val1)+2):
            sheet1.cell(row=i, column=15).value = val1[i-3]
            sheet1.cell(row=i, column=16).value = val_1[i-3]
            sheet1.cell(row=i, column=17).value = val2[i-3]
            sheet1.cell(row=i, column=18).value = val_2[i-3]
            sheet1.cell(row=i, column=19).value = val3[i-3]
            sheet1.cell(row=i, column=20).value = val_3[i-3]
            sheet1.cell(row=i, column=21).value = val4[i-3]
            sheet1.cell(row=i, column=22).value = val_4[i-3]
            for j in range(15, 23):
                sheet1.cell(row=i, column=j).border = thin_border
            i = i+1

        sheet1.cell(row=i+3, column=29).value = "Octant ID"
        sheet1.cell(row=i+3, column=30).value = "Octant Name"
        sheet1.cell(row=i+3, column=31).value = "Count of Rank 1 Mod Values"
        for col in range(29, 32):
            sheet1.cell(row=i+3, column=col).border = thin_border
        k = 0
        j = i+4
        while (k < len(oct_list)):
            sheet1.cell(row=j, column=29).value = oct_list[k]
            sheet1.cell(row=j, column=29).border = thin_border
            sheet1.cell(
                row=j, column=30).value = octant_name_id_mapping[oct_list[k]]
            sheet1.cell(row=j, column=30).border = thin_border
            sheet1.cell(row=j, column=31).value = count_Rank1_ModValue.count(
                int(oct_list[k]))
            sheet1.cell(row=j, column=31).border = thin_border
            j = j+1
            k = k+1

        overall_list = [val1[0], val_1[0], val2[0],
                        val_2[0], val3[0], val_3[0], val4[0], val_4[0]]
        overall_list_sorted = [val1[0], val_1[0], val2[0],
                            val_2[0], val3[0], val_3[0], val4[0], val_4[0]]
        overall_list_sorted.sort(reverse=True)
        # finding and storing rank for overall count
        sheet1.cell(row=3, column=23).value = overall_list_sorted.index(val1[0])+1
        sheet1.cell(row=3, column=24).value = overall_list_sorted.index(val_1[0])+1
        sheet1.cell(row=3, column=25).value = overall_list_sorted.index(val2[0])+1
        sheet1.cell(row=3, column=26).value = overall_list_sorted.index(val_2[0])+1
        sheet1.cell(row=3, column=27).value = overall_list_sorted.index(val3[0])+1
        sheet1.cell(row=3, column=28).value = overall_list_sorted.index(val_3[0])+1
        sheet1.cell(row=3, column=29).value = overall_list_sorted.index(val4[0])+1
        sheet1.cell(row=3, column=30).value = overall_list_sorted.index(val_4[0])+1
        for i in range(23, 31):
            sheet1.cell(row=3, column=i).border = thin_border
            if sheet1.cell(row=3, column=i).value == 1:
                sheet1.cell(row=3, column=i).fill = PatternFill(
                    start_color='FFD970', end_color='FFD970', fill_type="solid")  # to fill color to the cell

        # finding and storing rank1 Octant ID and rank1 Octant Name for overall count
        if (overall_list_sorted.index(val1[0])+1 == 1):
            sheet1.cell(row=3, column=31).value = 1
            sheet1.cell(row=3, column=32).value = octant_name_id_mapping["1"]
        elif (overall_list_sorted.index(val_1[0])+1 == 1):
            sheet1.cell(row=3, column=31).value = -1
            sheet1.cell(row=3, column=32).value = octant_name_id_mapping["-1"]
        elif (overall_list_sorted.index(val2[0])+1 == 1):
            sheet1.cell(row=3, column=31).value = 2
            sheet1.cell(row=3, column=32).value = octant_name_id_mapping["2"]
        elif (overall_list_sorted.index(val_2[0])+1 == 1):
            sheet1.cell(row=3, column=31).value = -2
            sheet1.cell(row=3, column=32).value = octant_name_id_mapping["-2"]
        elif (overall_list_sorted.index(val3[0])+1 == 1):
            sheet1.cell(row=3, column=31).value = 3
            sheet1.cell(row=3, column=32).value = octant_name_id_mapping["3"]
        elif (overall_list_sorted.index(val_3[0])+1 == 1):
            sheet1.cell(row=3, column=31).value = -3
            sheet1.cell(row=3, column=32).value = octant_name_id_mapping["-3"]
        elif (overall_list_sorted.index(val4[0])+1 == 1):
            sheet1.cell(row=3, column=31).value = 4
            sheet1.cell(row=3, column=32).value = octant_name_id_mapping["4"]
        elif (overall_list_sorted.index(val_4[0])+1 == 1):
            sheet1.cell(row=3, column=31).value = -4
            sheet1.cell(row=3, column=32).value = octant_name_id_mapping["-4"]

        sheet1.cell(row=3, column=31).border = thin_border
        sheet1.cell(row=3, column=32).border = thin_border

        sheet1.cell(row=1, column=35).value = "Overall Transition Count"
        sheet1.cell(row=2, column=36).value = "To"
        # creating a list to store octant ids

        oct_list = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
        j = 36
        i = 0
        while (j < 44 and i < len(oct_list)):
            sheet1.cell(row=3, column=j).value = oct_list[i]
            sheet1.cell(row=3, column=j).border = thin_border
            j = j+1
            i = i+1

        sheet1.cell(row=3, column=35).value = "Count"
        sheet1.cell(row=3, column=35).border = thin_border
        sheet1.cell(row=4, column=34).value = "From"

        k = 4
        i = 0
        while (k < 12 and i < len(oct_list)):
            sheet1.cell(row=k, column=35).value = oct_list[i]
            sheet1.cell(row=k, column=35).border = thin_border
            k = k+1
            i = i+1

        # variables for storing overall transition count
        r1c1 = r1c2 = r1c3 = r1c4 = r1c5 = r1c6 = r1c7 = r1c8 = 0
        r2c1 = r2c2 = r2c3 = r2c4 = r2c5 = r2c6 = r2c7 = r2c8 = 0
        r3c1 = r3c2 = r3c3 = r3c4 = r3c5 = r3c6 = r3c7 = r3c8 = 0
        r4c1 = r4c2 = r4c3 = r4c4 = r4c5 = r4c6 = r4c7 = r4c8 = 0
        r5c1 = r5c2 = r5c3 = r5c4 = r5c5 = r5c6 = r5c7 = r5c8 = 0
        r6c1 = r6c2 = r6c3 = r6c4 = r6c5 = r6c6 = r6c7 = r6c8 = 0
        r7c1 = r7c2 = r7c3 = r7c4 = r7c5 = r7c6 = r7c7 = r7c8 = 0
        r8c1 = r8c2 = r8c3 = r8c4 = r8c5 = r8c6 = r8c7 = r8c8 = 0
        # finding overall transition count
        i = 0
        while (i < len(octant_list)-1):
            if (octant_list[i] == 1):
                if (octant_list[i+1] == 1):
                    r1c1 += 1
                elif (octant_list[i+1] == -1):
                    r1c2 += 1
                elif (octant_list[i+1] == 2):
                    r1c3 += 1
                elif (octant_list[i+1] == -2):
                    r1c4 += 1
                elif (octant_list[i+1] == 3):
                    r1c5 += 1
                elif (octant_list[i+1] == -3):
                    r1c6 += 1
                elif (octant_list[i+1] == 4):
                    r1c7 += 1
                elif (octant_list[i+1] == -4):
                    r1c8 += 1
                i += 1
            elif (octant_list[i] == -1):
                if (octant_list[i+1] == 1):
                    r2c1 += 1
                elif (octant_list[i+1] == -1):
                    r2c2 += 1
                elif (octant_list[i+1] == 2):
                    r2c3 += 1
                elif (octant_list[i+1] == -2):
                    r2c4 += 1
                elif (octant_list[i+1] == 3):
                    r2c5 += 1
                elif (octant_list[i+1] == -3):
                    r2c6 += 1
                elif (octant_list[i+1] == 4):
                    r2c7 += 1
                elif (octant_list[i+1] == -4):
                    r2c8 += 1
                i += 1
            elif (octant_list[i] == 2):
                if (octant_list[i+1] == 1):
                    r3c1 += 1
                elif (octant_list[i+1] == -1):
                    r3c2 += 1
                elif (octant_list[i+1] == 2):
                    r3c3 += 1
                elif (octant_list[i+1] == -2):
                    r3c4 += 1
                elif (octant_list[i+1] == 3):
                    r3c5 += 1
                elif (octant_list[i+1] == -3):
                    r3c6 += 1
                elif (octant_list[i+1] == 4):
                    r3c7 += 1
                elif (octant_list[i+1] == -4):
                    r3c8 += 1
                i += 1
            elif (octant_list[i] == -2):
                if (octant_list[i+1] == 1):
                    r4c1 += 1
                elif (octant_list[i+1] == -1):
                    r4c2 += 1
                elif (octant_list[i+1] == 2):
                    r4c3 += 1
                elif (octant_list[i+1] == -2):
                    r4c4 += 1
                elif (octant_list[i+1] == 3):
                    r4c5 += 1
                elif (octant_list[i+1] == -3):
                    r4c6 += 1
                elif (octant_list[i+1] == 4):
                    r4c7 += 1
                elif (octant_list[i+1] == -4):
                    r4c8 += 1
                i += 1
            elif (octant_list[i] == 3):
                if (octant_list[i+1] == 1):
                    r5c1 += 1
                elif (octant_list[i+1] == -1):
                    r5c2 += 1
                elif (octant_list[i+1] == 2):
                    r5c3 += 1
                elif (octant_list[i+1] == -2):
                    r5c4 += 1
                elif (octant_list[i+1] == 3):
                    r5c5 += 1
                elif (octant_list[i+1] == -3):
                    r5c6 += 1
                elif (octant_list[i+1] == 4):
                    r5c7 += 1
                elif (octant_list[i+1] == -4):
                    r5c8 += 1
                i += 1
            elif (octant_list[i] == -3):
                if (octant_list[i+1] == 1):
                    r6c1 += 1
                elif (octant_list[i+1] == -1):
                    r6c2 += 1
                elif (octant_list[i+1] == 2):
                    r6c3 += 1
                elif (octant_list[i+1] == -2):
                    r6c4 += 1
                elif (octant_list[i+1] == 3):
                    r6c5 += 1
                elif (octant_list[i+1] == -3):
                    r6c6 += 1
                elif (octant_list[i+1] == 4):
                    r6c7 += 1
                elif (octant_list[i+1] == -4):
                    r6c8 += 1
                i += 1
            elif (octant_list[i] == 4):
                if (octant_list[i+1] == 1):
                    r7c1 += 1
                elif (octant_list[i+1] == -1):
                    r7c2 += 1
                elif (octant_list[i+1] == 2):
                    r7c3 += 1
                elif (octant_list[i+1] == -2):
                    r7c4 += 1
                elif (octant_list[i+1] == 3):
                    r7c5 += 1
                elif (octant_list[i+1] == -3):
                    r7c6 += 1
                elif (octant_list[i+1] == 4):
                    r7c7 += 1
                elif (octant_list[i+1] == -4):
                    r7c8 += 1
                i += 1
            elif (octant_list[i] == -4):
                if (octant_list[i+1] == 1):
                    r8c1 += 1
                elif (octant_list[i+1] == -1):
                    r8c2 += 1
                elif (octant_list[i+1] == 2):
                    r8c3 += 1
                elif (octant_list[i+1] == -2):
                    r8c4 += 1
                elif (octant_list[i+1] == 3):
                    r8c5 += 1
                elif (octant_list[i+1] == -3):
                    r8c6 += 1
                elif (octant_list[i+1] == 4):
                    r8c7 += 1
                elif (octant_list[i+1] == -4):
                    r8c8 += 1
                i += 1
        # lists for storing overall transition count
        trow1 = [r1c1, r2c1, r3c1, r4c1, r5c1, r6c1, r7c1, r8c1]
        trow2 = [r1c2, r2c2, r3c2, r4c2, r5c2, r6c2, r7c2, r8c2]
        trow3 = [r1c3, r2c3, r3c3, r4c3, r5c3, r6c3, r7c3, r8c3]
        trow4 = [r1c4, r2c4, r3c4, r4c4, r5c4, r6c4, r7c4, r8c4]
        trow5 = [r1c5, r2c5, r3c5, r4c5, r5c5, r6c5, r7c5, r8c5]
        trow6 = [r1c6, r2c6, r3c6, r4c6, r5c6, r6c6, r7c6, r8c6]
        trow7 = [r1c7, r2c7, r3c7, r4c7, r5c7, r6c7, r7c7, r8c7]
        trow8 = [r1c8, r2c8, r3c8, r4c8, r5c8, r6c8, r7c8, r8c8]

        # inserting overall transition count in sheet
        i = 0
        j = 4
        while (j <= 11 and i < len(trow1)):
            sheet1.cell(row=j, column=36).value = trow1[i]
            if i == 0:
                sheet1.cell(row=j, column=36).fill = PatternFill(
                    start_color='FFD970', end_color='FFD970', fill_type="solid")
            sheet1.cell(row=j, column=37).value = trow2[i]
            if i == 1:
                sheet1.cell(row=j, column=37).fill = PatternFill(
                    start_color='FFD970', end_color='FFD970', fill_type="solid")
            sheet1.cell(row=j, column=38).value = trow3[i]
            if i == 2:
                sheet1.cell(row=j, column=38).fill = PatternFill(
                    start_color='FFD970', end_color='FFD970', fill_type="solid")
            sheet1.cell(row=j, column=39).value = trow4[i]
            if i == 3:
                sheet1.cell(row=j, column=39).fill = PatternFill(
                    start_color='FFD970', end_color='FFD970', fill_type="solid")
            sheet1.cell(row=j, column=40).value = trow5[i]
            if i == 4:
                sheet1.cell(row=j, column=40).fill = PatternFill(
                    start_color='FFD970', end_color='FFD970', fill_type="solid")
            sheet1.cell(row=j, column=41).value = trow6[i]
            if i == 5:
                sheet1.cell(row=j, column=41).fill = PatternFill(
                    start_color='FFD970', end_color='FFD970', fill_type="solid")
            sheet1.cell(row=j, column=42).value = trow7[i]
            if i == 6:
                sheet1.cell(row=j, column=42).fill = PatternFill(
                    start_color='FFD970', end_color='FFD970', fill_type="solid")
            sheet1.cell(row=j, column=43).value = trow8[i]
            if i == 7:
                sheet1.cell(row=j, column=43).fill = PatternFill(
                    start_color='FFD970', end_color='FFD970', fill_type="solid")
            for col in range(36, 44):
                sheet1.cell(row=j, column=col).border = thin_border
            i = i+1
            j = j+1

        # funtion definition for finding mod transition count

        def transition_identification(mod):
            x = 0
            i = 0
            p = 1
            interval = mod

            while (x < n/interval):
                l = 36
                j = 0
                while (l < 44 and j < len(oct_list)):
                    sheet1.cell(row=p*13+2, column=l).value = oct_list[j]
                    sheet1.cell(row=p*13+2, column=l).border = thin_border
                    l = l+1
                    j = j+1
                sheet1.cell(row=p*13,
                            column=35).value = "Mod Transition Count"
                sheet1.cell(row=1+p*13,
                            column=35).value = str(interval*x)+"-"+str(interval*(x+1)-1)
                sheet1.cell(row=1+p*13, column=36).value = "To"
                sheet1.cell(row=2+p*13, column=35).value = "Count"
                sheet1.cell(row=2+p*13, column=35).border = thin_border
                sheet1.cell(row=3+p*13, column=34).value = "From"

                m = 3+p*13
                q = 0
                while (m < 11+p*13 and q < len(oct_list)):
                    sheet1.cell(row=m, column=35).value = oct_list[q]
                    sheet1.cell(row=m, column=35).border = thin_border
                    m = m+1
                    q = q+1
                # variables for storing mod transition count
                r1c1 = r1c2 = r1c3 = r1c4 = r1c5 = r1c6 = r1c7 = r1c8 = 0
                r2c1 = r2c2 = r2c3 = r2c4 = r2c5 = r2c6 = r2c7 = r2c8 = 0
                r3c1 = r3c2 = r3c3 = r3c4 = r3c5 = r3c6 = r3c7 = r3c8 = 0
                r4c1 = r4c2 = r4c3 = r4c4 = r4c5 = r4c6 = r4c7 = r4c8 = 0
                r5c1 = r5c2 = r5c3 = r5c4 = r5c5 = r5c6 = r5c7 = r5c8 = 0
                r6c1 = r6c2 = r6c3 = r6c4 = r6c5 = r6c6 = r6c7 = r6c8 = 0
                r7c1 = r7c2 = r7c3 = r7c4 = r7c5 = r7c6 = r7c7 = r7c8 = 0
                r8c1 = r8c2 = r8c3 = r8c4 = r8c5 = r8c6 = r8c7 = r8c8 = 0

                # finding  transition count in interval
                while (i < mod and i < len(octant_list)-1):
                    if (octant_list[i] == 1):
                        if (octant_list[i+1] == 1):
                            r1c1 += 1
                        elif (octant_list[i+1] == -1):
                            r1c2 += 1
                        elif (octant_list[i+1] == 2):
                            r1c3 += 1
                        elif (octant_list[i+1] == -2):
                            r1c4 += 1
                        elif (octant_list[i+1] == 3):
                            r1c5 += 1
                        elif (octant_list[i+1] == -3):
                            r1c6 += 1
                        elif (octant_list[i+1] == 4):
                            r1c7 += 1
                        elif (octant_list[i+1] == -4):
                            r1c8 += 1

                    elif (octant_list[i] == -1):
                        if (octant_list[i+1] == 1):
                            r2c1 += 1
                        elif (octant_list[i+1] == -1):
                            r2c2 += 1
                        elif (octant_list[i+1] == 2):
                            r2c3 += 1
                        elif (octant_list[i+1] == -2):
                            r2c4 += 1
                        elif (octant_list[i+1] == 3):
                            r2c5 += 1
                        elif (octant_list[i+1] == -3):
                            r2c6 += 1
                        elif (octant_list[i+1] == 4):
                            r2c7 += 1
                        elif (octant_list[i+1] == -4):
                            r2c8 += 1

                    elif (octant_list[i] == 2):
                        if (octant_list[i+1] == 1):
                            r3c1 += 1
                        elif (octant_list[i+1] == -1):
                            r3c2 += 1
                        elif (octant_list[i+1] == 2):
                            r3c3 += 1
                        elif (octant_list[i+1] == -2):
                            r3c4 += 1
                        elif (octant_list[i+1] == 3):
                            r3c5 += 1
                        elif (octant_list[i+1] == -3):
                            r3c6 += 1
                        elif (octant_list[i+1] == 4):
                            r3c7 += 1
                        elif (octant_list[i+1] == -4):
                            r3c8 += 1

                    elif (octant_list[i] == -2):
                        if (octant_list[i+1] == 1):
                            r4c1 += 1
                        elif (octant_list[i+1] == -1):
                            r4c2 += 1
                        elif (octant_list[i+1] == 2):
                            r4c3 += 1
                        elif (octant_list[i+1] == -2):
                            r4c4 += 1
                        elif (octant_list[i+1] == 3):
                            r4c5 += 1
                        elif (octant_list[i+1] == -3):
                            r4c6 += 1
                        elif (octant_list[i+1] == 4):
                            r4c7 += 1
                        elif (octant_list[i+1] == -4):
                            r4c8 += 1

                    elif (octant_list[i] == 3):
                        if (octant_list[i+1] == 1):
                            r5c1 += 1
                        elif (octant_list[i+1] == -1):
                            r5c2 += 1
                        elif (octant_list[i+1] == 2):
                            r5c3 += 1
                        elif (octant_list[i+1] == -2):
                            r5c4 += 1
                        elif (octant_list[i+1] == 3):
                            r5c5 += 1
                        elif (octant_list[i+1] == -3):
                            r5c6 += 1
                        elif (octant_list[i+1] == 4):
                            r5c7 += 1
                        elif (octant_list[i+1] == -4):
                            r5c8 += 1

                    elif (octant_list[i] == -3):
                        if (octant_list[i+1] == 1):
                            r6c1 += 1
                        elif (octant_list[i+1] == -1):
                            r6c2 += 1
                        elif (octant_list[i+1] == 2):
                            r6c3 += 1
                        elif (octant_list[i+1] == -2):
                            r6c4 += 1
                        elif (octant_list[i+1] == 3):
                            r6c5 += 1
                        elif (octant_list[i+1] == -3):
                            r6c6 += 1
                        elif (octant_list[i+1] == 4):
                            r6c7 += 1
                        elif (octant_list[i+1] == -4):
                            r6c8 += 1

                    elif (octant_list[i] == 4):
                        if (octant_list[i+1] == 1):
                            r7c1 += 1
                        elif (octant_list[i+1] == -1):
                            r7c2 += 1
                        elif (octant_list[i+1] == 2):
                            r7c3 += 1
                        elif (octant_list[i+1] == -2):
                            r7c4 += 1
                        elif (octant_list[i+1] == 3):
                            r7c5 += 1
                        elif (octant_list[i+1] == -3):
                            r7c6 += 1
                        elif (octant_list[i+1] == 4):
                            r7c7 += 1
                        elif (octant_list[i+1] == -4):
                            r7c8 += 1

                    elif (octant_list[i] == -4):
                        if (octant_list[i+1] == 1):
                            r8c1 += 1
                        elif (octant_list[i+1] == -1):
                            r8c2 += 1
                        elif (octant_list[i+1] == 2):
                            r8c3 += 1
                        elif (octant_list[i+1] == -2):
                            r8c4 += 1
                        elif (octant_list[i+1] == 3):
                            r8c5 += 1
                        elif (octant_list[i+1] == -3):
                            r8c6 += 1
                        elif (octant_list[i+1] == 4):
                            r8c7 += 1
                        elif (octant_list[i+1] == -4):
                            r8c8 += 1
                    i += 1
                # storing mod transition count in lists
                trow1 = [r1c1, r2c1, r3c1, r4c1, r5c1, r6c1, r7c1, r8c1]
                trow2 = [r1c2, r2c2, r3c2, r4c2, r5c2, r6c2, r7c2, r8c2]
                trow3 = [r1c3, r2c3, r3c3, r4c3, r5c3, r6c3, r7c3, r8c3]
                trow4 = [r1c4, r2c4, r3c4, r4c4, r5c4, r6c4, r7c4, r8c4]
                trow5 = [r1c5, r2c5, r3c5, r4c5, r5c5, r6c5, r7c5, r8c5]
                trow6 = [r1c6, r2c6, r3c6, r4c6, r5c6, r6c6, r7c6, r8c6]
                trow7 = [r1c7, r2c7, r3c7, r4c7, r5c7, r6c7, r7c7, r8c7]
                trow8 = [r1c8, r2c8, r3c8, r4c8, r5c8, r6c8, r7c8, r8c8]
                # inserting mod transition count in sheet
                y = 0
                z = p*13+3
                while (z <= 10+p*13 and y < len(trow1)):
                    sheet1.cell(row=z, column=36).value = trow1[y]
                    if y == 0:
                        sheet1.cell(row=z, column=36).fill = PatternFill(
                            start_color='FFD970', end_color='FFD970', fill_type="solid")
                    sheet1.cell(row=z, column=37).value = trow2[y]
                    if y == 1:
                        sheet1.cell(row=z, column=37).fill = PatternFill(
                            start_color='FFD970', end_color='FFD970', fill_type="solid")
                    sheet1.cell(row=z, column=38).value = trow3[y]
                    if y == 2:
                        sheet1.cell(row=z, column=38).fill = PatternFill(
                            start_color='FFD970', end_color='FFD970', fill_type="solid")
                    sheet1.cell(row=z, column=39).value = trow4[y]
                    if y == 3:
                        sheet1.cell(row=z, column=39).fill = PatternFill(
                            start_color='FFD970', end_color='FFD970', fill_type="solid")
                    sheet1.cell(row=z, column=40).value = trow5[y]
                    if y == 4:
                        sheet1.cell(row=z, column=40).fill = PatternFill(
                            start_color='FFD970', end_color='FFD970', fill_type="solid")
                    sheet1.cell(row=z, column=41).value = trow6[y]
                    if y == 5:
                        sheet1.cell(row=z, column=41).fill = PatternFill(
                            start_color='FFD970', end_color='FFD970', fill_type="solid")
                    sheet1.cell(row=z, column=42).value = trow7[y]
                    if y == 6:
                        sheet1.cell(row=z, column=42).fill = PatternFill(
                            start_color='FFD970', end_color='FFD970', fill_type="solid")
                    sheet1.cell(row=z, column=43).value = trow8[y]
                    if y == 7:
                        sheet1.cell(row=z, column=43).fill = PatternFill(
                            start_color='FFD970', end_color='FFD970', fill_type="solid")
                    for col in range(36, 44):
                        sheet1.cell(row=z, column=col).border = thin_border
                    y = y+1
                    z = z+1
                mod = mod+interval
                p = p+1
                x = x+1

        # value of mod in line no-141
        # funtion calling
        # funtion for finding mod transition count
        transition_identification(mod)

        def octant_longest_subsequence_count():
            # variables for counting the subsequence length
            count1 = count_1 = count2 = count_2 = count3 = count_3 = count4 = count_4 = 1

            for i in range(len(octant_list)-1):
                # appending the subsequence length stored in count1 variable in list1
                if (octant_list[i] == 1):
                    if (octant_list[i+1] == 1):
                        count1 += 1
                    else:
                        list1.append(count1)
                        count1 = 1
                # appending the subsequence length stored in count_1 variable in list_1
                elif (octant_list[i] == -1):
                    if (octant_list[i+1] == -1):
                        count_1 += 1
                    else:
                        list_1.append(count_1)
                        count_1 = 1

                # appending the subsequence length stored in count2 variable in list2
                elif (octant_list[i] == 2):
                    if (octant_list[i+1] == 2):
                        count2 += 1
                    else:
                        list2.append(count2)
                        count2 = 1
                # appending the subsequence length stored in count_2 variable in list_2
                elif (octant_list[i] == -2):
                    if (octant_list[i+1] == -2):
                        count_2 += 1
                    else:
                        list_2.append(count_2)
                        count_2 = 1

                # appending the subsequence length stored in count3 variable in list3
                elif (octant_list[i] == 3):
                    if (octant_list[i+1] == 3):
                        count3 += 1
                    else:
                        list3.append(count3)
                        count3 = 1
                # appending the subsequence length stored in count_3 variable in list_3
                elif (octant_list[i] == -3):
                    if (octant_list[i+1] == -3):
                        count_3 += 1
                    else:
                        list_3.append(count_3)
                        count_3 = 1

                # appending the subsequence length stored in count4 variable in list4
                elif (octant_list[i] == 4):
                    if (octant_list[i+1] == 4):
                        count4 += 1
                    else:
                        list4.append(count4)
                        count4 = 1
                # appending the subsequence length stored in count_4 variable in list_4
                elif (octant_list[i] == -4):
                    if (octant_list[i+1] == -4):
                        count_4 += 1
                    else:
                        list_4.append(count_4)
                        count_4 = 1

        sheet1.cell(row=3, column=45).value = "Octant"
        sheet1.cell(row=3, column=45).border = thin_border
        sheet1.cell(row=3, column=46).value = "Longest Subsequence Length"
        sheet1.cell(row=3, column=46).border = thin_border
        sheet1.cell(row=3, column=47).value = "Count"
        sheet1.cell(row=3, column=47).border = thin_border

        oct_list = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
        i = 0
        while (i < len(oct_list)):
            sheet1.cell(row=i+4, column=45).value = oct_list[i]
            sheet1.cell(row=i+4, column=45).border = thin_border
            i = i+1

        # lists for storing all subsequence length
        list1 = []
        list_1 = []
        list2 = []
        list_2 = []
        list3 = []
        list_3 = []
        list4 = []
        list_4 = []

        # funtion calling
        octant_longest_subsequence_count()

        # inserting maximum value of lists(list1, list_1, .....list4, list_4),which will be longest subsequence length) in sheet

        sheet1.cell(row=4, column=46).value = max(list1)
        sheet1.cell(row=5, column=46).value = max(list_1)
        sheet1.cell(row=6, column=46).value = max(list2)
        sheet1.cell(row=7, column=46).value = max(list_2)
        sheet1.cell(row=8, column=46).value = max(list3)
        sheet1.cell(row=9, column=46).value = max(list_3)
        sheet1.cell(row=10, column=46).value = max(list4)
        sheet1.cell(row=11, column=46).value = max(list_4)
        for r in range(4, 12):
            sheet1.cell(row=r, column=46).border = thin_border

        # inserting count of maximum value of lists in sheet
        sheet1.cell(row=4, column=47).value = list1.count(max(list1))
        sheet1.cell(row=5, column=47).value = list_1.count(max(list_1))
        sheet1.cell(row=6, column=47).value = list2.count(max(list2))
        sheet1.cell(row=7, column=47).value = list_2.count(max(list_2))
        sheet1.cell(row=8, column=47).value = list3.count(max(list3))
        sheet1.cell(row=9, column=47).value = list_3.count(max(list_3))
        sheet1.cell(row=10, column=47).value = list4.count(max(list4))
        sheet1.cell(row=11, column=47).value = list_4.count(max(list_4))

        for r in range(4, 12):
            sheet1.cell(row=r, column=47).border = thin_border

        def octant_longest_subsequence_count_with_range():
            # variables for counting the subsequence length
            count1 = count_1 = count2 = count_2 = count3 = count_3 = count4 = count_4 = 1

            for i in range(len(octant_list)-1):
                # storing the time inteval of longest subsequence for all 8 octants
                if (octant_list[i] == 1):
                    if (octant_list[i+1] == 1):
                        count1 += 1
                    else:
                        if (count1 == max(list1)):
                            time1.append([time[i-count1+1], time[i]])
                        count1 = 1

                elif (octant_list[i] == -1):
                    if (octant_list[i+1] == -1):
                        count_1 += 1
                    else:
                        if (count_1 == max(list_1)):
                            time_1.append([time[i-count_1+1], time[i]])
                        count_1 = 1

                elif (octant_list[i] == 2):
                    if (octant_list[i+1] == 2):
                        count2 += 1
                    else:
                        if (count2 == max(list2)):
                            time2.append([time[i-count2+1], time[i]])
                        count2 = 1

                elif (octant_list[i] == -2):
                    if (octant_list[i+1] == -2):
                        count_2 += 1
                    else:
                        if (count_2 == max(list_2)):
                            time_2.append([time[i-count_2+1], time[i]])
                        count_2 = 1

                elif (octant_list[i] == 3):
                    if (octant_list[i+1] == 3):
                        count3 += 1
                    else:
                        if (count3 == max(list3)):
                            time3.append([time[i-count3+1], time[i]])
                        count3 = 1

                elif (octant_list[i] == -3):
                    if (octant_list[i+1] == -3):
                        count_3 += 1
                    else:
                        if (count_3 == max(list_3)):
                            time_3.append([time[i-count_3+1], time[i]])
                        count_3 = 1

                elif (octant_list[i] == 4):
                    if (octant_list[i+1] == 4):
                        count4 += 1
                    else:
                        if (count4 == max(list4)):
                            time4.append([time[i-count4+1], time[i]])
                        count4 = 1

                elif (octant_list[i] == -4):
                    if (octant_list[i+1] == -4):
                        count_4 += 1
                    else:
                        if (count_4 == max(list_4)):
                            time_4.append([time[i-count_4+1], time[i]])
                        count_4 = 1

        sheet1.cell(row=1, column=49).value = "Longest Subsequence Length with Range"
        sheet1.cell(row=3, column=49).value = "Octant"
        sheet1.cell(row=3, column=49).border = thin_border
        sheet1.cell(row=3, column=50).value = "Longest Subsequence Length"
        sheet1.cell(row=3, column=50).border = thin_border
        sheet1.cell(row=3, column=51).value = "Count"
        sheet1.cell(row=3, column=51).border = thin_border
        oct_list = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
        i = 0
        while (i < len(oct_list)):
            sheet1.cell(row=i+3, column=49).value = oct_list[i]
            sheet1.cell(row=i+3, column=49).border = thin_border
            i = i+1

        # lists for storing all subsequence length
        # list1, list_1, list2, list_2, list3, list_3, list4, list_4 = (
        #     [0] for i in range(8))
        # lists for storing time interval of longest subsequence
        time1, time_1, time2, time_2, time3, time_3, time4, time_4 = (
            [] for i in range(8))

        # funtion calling
        octant_longest_subsequence_count_with_range()

        # storing the values to be inserted in column-17, 18, 19 in lists
        col_17 = ["+1", "Time", " ", "-1", "Time", " ", "+2", "Time", " ", "-2",
                "Time", " ", "+3", "Time", " ", "-3", "Time", " ", "+4", "Time", " ", "-4", "Time", " "]

        col_18 = [max(list1), "From", time1[0][0], max(list_1), "From", time_1[0][0], max(list2), "From", time2[0][0], max(list_2), "From", time_2[0][0], max(
            list3), "From", time3[0][0], max(list_3), "From", time_3[0][0], max(list4), "From", time4[0][0], max(list_4), "From", time_4[0][0]]

        # print(col_18)

        col_19 = [list1.count(max(list1)), "To", time1[0][1], list_1.count(max(list_1)), "To", time_1[0][1], list2.count(max(list2)), "To", time2[0][1], list_2.count(max(list_2)), "To", time_2[0][1], list3.count(
            max(list3)), "To", time3[0][1], list_3.count(max(list_3)), "To", time_3[0][1], list4.count(max(list4)), "To", time4[0][1], list_4.count(max(list_4)), "To", time_4[0][1]]

        # inserting time intervals in sheet1
        for i in range(3, len(col_17)+3):
            sheet1.cell(row=i, column=49).value = col_17[i-3]
            sheet1.cell(row=i, column=49).border = thin_border
            sheet1.cell(row=i, column=50).value = col_18[i-3]
            sheet1.cell(row=i, column=50).border = thin_border
            sheet1.cell(row=i, column=51).value = col_19[i-3]
            sheet1.cell(row=i, column=51).border = thin_border
        
        currentDateAndTime = datetime.now()
        DateAndTime = str(currentDateAndTime.year) + '-' + str(currentDateAndTime.month) + '-' + str(currentDateAndTime.day) + '-' + str(currentDateAndTime.hour) + '-' + str(currentDateAndTime.minute) + '-' + str(currentDateAndTime.second)
        if selected == "Upload File":
            wb.save('output/' + str(file.name[:-5]) + '_' + str(mod) + '_' + str(DateAndTime) + '.xlsx')
        else:
            wb.save('output/' + str(file[:-5]) + '_' + str(mod) + '_' + str(DateAndTime) + '.xlsx')

if selected == "Upload File":
    uploaded_files = st.file_uploader("Upload a file or multiple files", accept_multiple_files=True)
    st.write("Number of files uploaded : ", len(uploaded_files))
    st.write("Uploaded files are:")
    for uploaded_file in uploaded_files:
        st.write(uploaded_file.name)
    mod_value=int(st.number_input("Enter the Mod Value"))
    st.write("The value of Mod is ", mod_value)
    st.write("Click Compute for computation")
    if st.button("Compute"):
        proj_octant_gui(uploaded_files, mod_value)
        st.write("Computation Done")

if selected == "Give Path":
    file_path=st.text_input("Enter the path of directory")
    st.write("Path is :", file_path)
    mod_value=int(st.number_input("Enter the Mod Value"))
    st.write("The value of Mod is ", mod_value)
    st.write("Click Compute for computation")
    if st.button("Compute"):
        proj_octant_gui(data, mod_value)
        st.write("Computation Done")

# This shall be the last lines of the code.
# end_time = datetime.now()
# print('Duration of Program Execution: {}'.format(end_time - start_time))

    
